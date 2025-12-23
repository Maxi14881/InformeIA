import base64
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

def crear_hoja_resumen_caracteristica(wb, df_caracteristica, caracteristica, version_agente, numero):
    """Crear una hoja de resumen para una característica específica"""
    pivot_car = pd.pivot_table(
        df_caracteristica,
        index='Id del caso de prueba',
        columns='Código de ejecución',
        values='Resultado',
        aggfunc='mean',
        fill_value=0
    )
    
    # Usar nomenclatura simplificada: Resumen_Ctca_1, Resumen_Ctca_2, etc.
    nombre_hoja = f"Resumen_Ctca_{numero}"
    
    # Crear nueva hoja
    ws_car = wb.create_sheet(nombre_hoja)
    
    # Escribir encabezado
    encabezado_data = [
        ['', ''],
        ['Informe de pruebas IA', ''],
        ['', ''],
        [f'ANÁLISIS GENERAL DEL ASISTENTE - {caracteristica}', ''],
        ['Versión del asistente', version_agente],
        ['Fecha de ejecución', datetime.today().strftime("%d/%m/%Y")],
        ['Cantidad de ejecuciones', df_caracteristica["Código de ejecución"].nunique()],
        ['', ''],
        ['CRITERIO ACEPTACIÓN - REGLA 80 - 80', ''],
        ['Total de casos analizados', ''],
        ['Total de casos con efectividad >= 80%', ''],
        ['% casos > 80%', ''],
        ['', ''],
        ['CRITERIO ACEPTACIÓN - NO FAILS', ''],
        ['Total de casos sin casos de éxito', '']
    ]
    
    for row_idx, (col_a, col_b) in enumerate(encabezado_data, 1):
        ws_car.cell(row=row_idx, column=1, value=col_a)
        ws_car.cell(row=row_idx, column=2, value=col_b)
    
    # Escribir pivot table
    fila_inicio_pivot = 17
    ws_car.cell(row=fila_inicio_pivot, column=1, value="Id del caso de prueba")
    for col_idx, col_name in enumerate(pivot_car.columns, 2):
        ws_car.cell(row=fila_inicio_pivot, column=col_idx, value=col_name)
    ws_car.cell(row=fila_inicio_pivot, column=len(pivot_car.columns) + 2, value="Total Gral")
    
    for row_idx, (index, row_data) in enumerate(pivot_car.iterrows(), fila_inicio_pivot + 1):
        ws_car.cell(row=row_idx, column=1, value=index)
        for col_idx, value in enumerate(row_data, 2):
            ws_car.cell(row=row_idx, column=col_idx, value=value)
    
    # Aplicar formato y estilos
    ws_car["A2"].font = Font(bold=True, size=20)
    ws_car.row_dimensions[2].height = 40
    ws_car.column_dimensions["A"].width = 40
    
    ws_car["A4"].font = Font(bold=True)
    ws_car["A9"].font = Font(bold=True)
    ws_car["A14"].font = Font(bold=True)
    
    # Cálculos dinámicos
    fila_inicio = fila_inicio_pivot + 1
    num_filas = pivot_car.shape[0]
    fila_fin = fila_inicio + num_filas - 1
    col_total = len(pivot_car.columns) + 2
    rango_g = f"{get_column_letter(col_total)}{fila_inicio}:{get_column_letter(col_total)}{fila_fin}"
    
    ws_car["B4"] = f'=IF(AND(B12>0.8,B15=0),"PASSED","FAILED")'
    ws_car["B10"] = f'=COUNTA({rango_g})'
    ws_car["B11"] = f'=COUNTIF({rango_g},">=80%")'
    ws_car["B12"] = '=B11/B10'
    ws_car["B12"].number_format = "0%"
    ws_car["B15"] = f'=COUNTIF({rango_g},"=0")'
    
    # Agregar fórmulas de promedio en columna Total Gral
    col_inicio = 2
    col_fin = col_inicio + pivot_car.shape[1] - 1
    
    for i in range(num_filas):
        fila_excel = fila_inicio + i
        col_letra_ini = get_column_letter(col_inicio)
        col_letra_fin = get_column_letter(col_fin)
        ws_car.cell(row=fila_excel, column=col_total, value=f"=AVERAGE({col_letra_ini}{fila_excel}:{col_letra_fin}{fila_excel})")
    
    # Fila de promedios
    fila_total = fila_inicio + num_filas
    ws_car.cell(row=fila_total, column=1, value="Promedio columna").font = Font(bold=True)
    for col in range(col_inicio, col_total + 1):
        col_letra = get_column_letter(col)
        ws_car.cell(row=fila_total, column=col, value=f"=AVERAGE({col_letra}{fila_inicio}:{col_letra}{fila_inicio + num_filas - 1})")
        ws_car.cell(row=fila_total, column=col).number_format = "0%"
    
    # Formatear como porcentajes
    for row in ws_car.iter_rows(min_row=fila_inicio, max_row=fila_total, min_col=col_inicio, max_col=col_total):
        for cell in row:
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = "0%"
    
    # Formato condicional
    ws_car.conditional_formatting.add(
        "B4",
        FormulaRule(formula=['B4="PASSED"'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
    )
    ws_car.conditional_formatting.add(
        "B4",
        FormulaRule(formula=['B4="FAILED"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    ws_car.conditional_formatting.add(
        "B12",
        FormulaRule(formula=['B12>=0.8'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
    )
    ws_car.conditional_formatting.add(
        "B12",
        FormulaRule(formula=['B12<0.8'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    ws_car.conditional_formatting.add(
        "B15",
        FormulaRule(formula=['B15=0'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
    )
    ws_car.conditional_formatting.add(
        "B15",
        FormulaRule(formula=['B15<>0'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    
    # Escala de colores para la tabla
    rango_color = f"{get_column_letter(col_inicio)}{fila_inicio}:{get_column_letter(col_total)}{fila_fin}"
    ws_car.conditional_formatting.add(
        rango_color,
        ColorScaleRule(
            start_type='num', start_value=0, start_color='FFFC7B7B',
            mid_type='num', mid_value=0.5, mid_color='FFFFFF00',
            end_type='num', end_value=1, end_color='FF63BE7B'
        )
    )
    
    # Bordes
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws_car.iter_rows(min_row=4, max_row=16, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    
    return ws_car

st.set_page_config(
    page_title="Exportador de Casos de Prueba",
    page_icon="📤",
)

# Convierte la imagen a base64
def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()

# URL de la imagen y enlace a YouTube
logo_path = "binit.png"
link = "https://binit.tech/"
logo_base64 = image_to_base64(logo_path)

# Mostrar imagen enlazada
st.markdown(
    f'<a href="{link}" target="_blank">'
    f'<img src="data:image/jpeg;base64,{logo_base64}" style="width:100%;"/>'
    '</a>',
    unsafe_allow_html=True
)

# CSS personalizado
st.markdown(
    """
    <style>
        .stApp {
            background-color: #E0DBDB;
        }
        padding: 10px 15px;
        border-radius: 8px;
        font-weight: bold;
        color: white; /* Color del texto en blanco */
        background-color: #333333; /* Gris oscuro para el fondo */
        border: none;
        cursor: pointer;
        transition: background-color 0.3s ease;
        }
    </style>
    """,
    unsafe_allow_html=True
)

st.title("Generador de Informe QAStudio")

uploaded_file = st.file_uploader("Sube tu archivo Excel de origen", type=["xlsx", "xls"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()
    
    # Mapeo de columnas inglés -> español
    column_mapping = {
        'Result': 'Resultado',
        'Test Case ID': 'Id del caso de prueba',
        'Execution Code': 'Código de ejecución',
        'Agent Version': 'Versión del agente',
        'Characteristic': 'Característica',
        'Test Case': 'Caso de prueba',
        'Expected Result': 'Resultado esperado'
    }
    
    # Detectar idioma y renombrar columnas si es necesario
    columnas_en_ingles = [col for col in column_mapping.keys() if col in df.columns]
    columnas_en_espanol = [col for col in column_mapping.values() if col in df.columns]
    
    if columnas_en_ingles:
        st.info(f"🌐 Detectadas columnas en inglés. Realizando traducción automática...")
        # Renombrar columnas de inglés a español
        df.rename(columns=column_mapping, inplace=True)
        st.success(f"✅ Columnas traducidas: {', '.join(columnas_en_ingles)}")
    elif columnas_en_espanol:
        st.info(f"🇪🇸 Columnas detectadas en español.")
    
    # Mostrar información de depuración sobre las columnas
    st.write("**Columnas después del mapeo:**")
    st.write(list(df.columns))
    
    # Verificar si existe la columna 'Resultado'
    if 'Resultado' not in df.columns:
        st.error(f"❌ No se encontró la columna 'Resultado' en el archivo.")
        st.error(f"📋 Columnas disponibles: {', '.join(df.columns)}")
        st.info("💡 Asegúrate de que tu archivo Excel contenga una columna llamada 'Resultado' o 'Result'")
        st.stop()
    
    df['Resultado'] = pd.to_numeric(df['Resultado'], errors='coerce')

    if 'Versión del agente' in df.columns:
        version_agente = str(df['Versión del agente'].max())
    else:
        version_agente = "N/A"

    # Verificar si existe columna Característica
    tiene_caracteristicas = 'Característica' in df.columns
    caracteristicas_unicas = []
    
    if tiene_caracteristicas:
        caracteristicas_unicas = df['Característica'].dropna().unique().tolist()
        st.info(f"📊 Detectadas {len(caracteristicas_unicas)} características: {', '.join(map(str, caracteristicas_unicas))}")
        if len(caracteristicas_unicas) > 1:
            st.info("Se generarán resúmenes individuales por cada característica.")
        else:
            st.info("Se generará un resumen general y resúmenes individuales por cada característica.")

    pivot = pd.pivot_table(
        df,
        index='Id del caso de prueba',
        columns='Código de ejecución',
        values='Resultado',
        aggfunc='mean',
        fill_value=0
    )

    output = BytesIO()
    wb = None
    
    # Solo generar hoja RESUMEN si hay 1 característica o sin características
    if not tiene_caracteristicas or len(caracteristicas_unicas) <= 1:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            encabezado = pd.DataFrame({
                'A': [
                    '', '', '', 
                    'ANÁLISIS GENERAL DEL ASISTENTE', 'Versión del asistente', 'Fecha de ejecución', 'Cantidad de ejecuciones', '',
                    'CRITERIO ACEPTACIÓN - REGLA 80 - 80', 'Total de casos analizados', 'Total de casos con efectividad >= 80%', '% casos > 80%',
                    '', 
                    'CRITERIO ACEPTACIÓN - NO FAILS', 'Total de casos sin casos de éxito'
                ],
                'B': ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
            })
            encabezado.to_excel(writer, index=False, header=False, startrow=0)
            pivot.to_excel(writer, startrow=16)
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        ws.title = "RESUMEN"
    else:
        # Para múltiples características, crear workbook vacío
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        wb.remove(ws)  # Remover hoja por defecto

    # Solo formatear RESUMEN si no hay múltiples características
    if not tiene_caracteristicas or len(caracteristicas_unicas) <= 1:
        ws["A2"] = "Informe de pruebas IA"
        ws["A2"].font = Font(bold=True, size=20)
        ws.row_dimensions[2].height = 40         # <-- Más alto el título
        ws.column_dimensions["A"].width = 40     # <-- Más ancha la columna A

        ws["A4"].font = Font(bold=True)
        ws["A9"].font = Font(bold=True)
        ws["A13"].font = Font(bold=True)

        ws["B4"] = '=IF(AND(B12>0.8,B15=0),"PASSED","FAILED")'
        ws["B5"] = version_agente
        ws["B6"] = datetime.today().strftime("%d/%m/%Y")
        ws["B7"] = df["Código de ejecución"].nunique()  # Cantidad de ejecuciones únicas

        fila_inicio = 18
        num_filas = pivot.shape[0]
        fila_fin = fila_inicio + num_filas - 1
        rango_g = f"G{fila_inicio}:G{fila_fin}"

        ws["B10"] = f'=COUNTA({rango_g})'
        ws["B11"] = f'=COUNTIF({rango_g},">=80%")'
        ws["B12"] = '=B11/B10'
        ws["B12"].number_format = "0%"
        ws["B15"] = f'=COUNTIF({rango_g},"=0")'

        ws["G17"] = "Total Gral"
        ws["G17"].font = Font(bold=True)

        col_inicio = 2
        col_fin = col_inicio + pivot.shape[1] - 1

        for i in range(num_filas):
            fila_excel = fila_inicio + i
            col_letra_ini = ws.cell(row=fila_excel, column=col_inicio).column_letter
            col_letra_fin = ws.cell(row=fila_excel, column=col_fin).column_letter
            ws[f"G{fila_excel}"] = f"=AVERAGE({col_letra_ini}{fila_excel}:{col_letra_fin}{fila_excel})"

        fila_total = fila_inicio + num_filas
        ws.cell(row=fila_total, column=1, value="Promedio columna").font = Font(bold=True)
        for col in range(col_inicio, col_fin + 2):
            col_letra = ws.cell(row=fila_inicio, column=col).column_letter
            ws.cell(row=fila_total, column=col, value=f"=AVERAGE({col_letra}{fila_inicio}:{col_letra}{fila_inicio + num_filas - 1})")
            ws.cell(row=fila_total, column=col).number_format = "0%"

        min_row = fila_inicio
        max_row = ws.max_row
        min_col = col_inicio
        max_col = col_fin + 1

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = "0%"

        ws.conditional_formatting.add(
            "B4",
            FormulaRule(formula=['B4="PASSED"'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            "B4",
            FormulaRule(formula=['B4="FAILED"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            "B12",
            FormulaRule(formula=['B12>=0.8'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            "B12",
            FormulaRule(formula=['B12<0.8'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            "B15",
            FormulaRule(formula=['B15=0'], fill=PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            "B15",
            FormulaRule(formula=['B15<>0'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
        )

        col_letra_ini = get_column_letter(min_col)
        col_letra_fin = get_column_letter(max_col)
        rango_color = f"{col_letra_ini}{min_row}:{col_letra_fin}{fila_fin}"

        ws.conditional_formatting.add(
            rango_color,
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FFFC7B7B',
                mid_type='num', mid_value=0.5, mid_color='FFFFFF00',
                end_type='num', end_value=1, end_color='FF63BE7B'
            )
        )

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=4, max_row=16, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

    # --- Generar hojas por característica si existen múltiples ---
    if tiene_caracteristicas and len(caracteristicas_unicas) > 1:
        st.info(f"🔄 Generando {len(caracteristicas_unicas)} hojas de resumen por característica...")
        st.info("ℹ️ Al tener múltiples características, se omite la hoja RESUMEN general.")
        for idx, caracteristica in enumerate(caracteristicas_unicas, 1):
            # Filtrar datos por característica
            df_caracteristica = df[df['Característica'] == caracteristica].copy()
            if not df_caracteristica.empty:
                crear_hoja_resumen_caracteristica(wb, df_caracteristica, caracteristica, version_agente, idx)
        st.success(f"✅ Hojas de resumen por característica generadas exitosamente.")
    elif tiene_caracteristicas and len(caracteristicas_unicas) == 1:
        st.info(f"ℹ️ Una sola característica detectada: {caracteristicas_unicas[0]}. Usando hoja RESUMEN general.")

    # --- Copiar hoja "Ejecuciones" si existe ---
    xls = pd.ExcelFile(uploaded_file)
    if "Ejecuciones" in xls.sheet_names:
        df_ejec = pd.read_excel(xls, sheet_name="Ejecuciones")
        ws_ejec = wb.create_sheet("EJECUCIONES")
        for i, col in enumerate(df_ejec.columns, 1):
            ws_ejec.cell(row=1, column=i, value=col)
        for row_idx, row in enumerate(df_ejec.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_ejec.cell(row=row_idx, column=col_idx, value=value)

    # --- Copiar hoja "CRITERIOS ACEPTACIÓN" ---
    columnas_criterios = ["Id del caso de prueba", "Caso de prueba", "Resultado esperado"]
    if all(col in df.columns for col in columnas_criterios):
        criterios_df = df[columnas_criterios].drop_duplicates().sort_values(by="Id del caso de prueba")
        ws_criterios = wb.create_sheet("CRITERIOS ACEPTACIÓN")
        for i, col in enumerate(criterios_df.columns, 1):
            ws_criterios.cell(row=1, column=i, value=col)
        for row_idx, row in enumerate(criterios_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_criterios.cell(row=row_idx, column=col_idx, value=value)

    # Guardar el archivo final
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)

    st.success("¡Archivo generado correctamente!")
    st.download_button(
        label="Descargar QAStudio Result.xlsx",
        data=output_final,
        file_name="QAStudio Result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
